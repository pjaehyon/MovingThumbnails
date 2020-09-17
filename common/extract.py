from openpyxl import load_workbook 
from bs4 import BeautifulSoup
import sys

def excel_data(excel_file, sheet_name, thumbnail_index, links_index, product_code_index):
    '''
    extract required column data.
    input: excel_file: Must be .xlsx // sheet_name: string // thumbnail_index, links_index, product_code_index: integer
    output: list of openpyxl cell objects
    '''
    try:
        wb = load_workbook(excel_file)
        sheet = wb[sheet_name]
        items = []
        
        for row in sheet.iter_rows():
            item = {}
            item['thumbnail'] = row[thumbnail_index]
            item['html_doc'] = row[links_index]
            item['product_code'] = row[product_code_index]
            items.append(item)

        del items[0]
    
    except:
        print('Format must be "xlsx". Convert to xlsx and try again.')
        sys.exit(1)

    return items

def cell_values(item):
    '''
    extract values of a cell for each column as string
    input: an openpyxl cell object
    output: tuple of 3 string values 
    '''
    thumbnail = item.get('thumbnail').value
    product_code = item.get('product_code').value
    img_links = links(item.get('html_doc').value)

    return thumbnail, product_code, img_links

def links(html_doc):
    '''
    extract img link for a cell
    input: html string
    output: list of links
    '''
    soup = BeautifulSoup(html_doc, 'html.parser')
    links = []
    
    for elem in soup.find_all('img'):
        links.append(elem.get('src'))
    
    return links